import pandas as pd
import numpy as np
import glob
import os
from scipy import stats
import matplotlib.pyplot as plt
import re
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment

class AdaptiveFreezingDetector:
    
    def __init__(self, animals_per_video=1, fps=30):
        self.animals_per_video = animals_per_video
        self.fps = fps
        self.sensitivity_levels = [0.5, 1.0, 1.5, 2.0, 2.5, 3.0, 3.5, 4.0, 4.5, 5.0]
        self.min_freezing_duration = 0.8
        self.max_freezing_duration = 15.0
        self.feature_weights = {
            'stability': 0.50,
            'relative_confidence': 0.30,
            'correlation': 0.20
        }
        print(f"ADAPTIVE FREEZING DETECTOR")
        print(f"Testing {len(self.sensitivity_levels)} sensitivity levels")
        print(f"Normalizing confidence relatively per video")
        print(f"Accounting for {animals_per_video} animals per video")
        print(f"MINIMUM BOUT DURATION: {self.min_freezing_duration}s")
    
    def load_data(self, csv_file):
        df = pd.read_csv(csv_file)
        available_bodyparts = []
        for col in df.columns:
            if col.endswith('_confidence') and not col.startswith('frame'):
                bp_name = col.replace('_confidence', '')
                available_bodyparts.append(bp_name)
        print(f"   Loaded {len(df)} frames, {len(available_bodyparts)} body parts")
        return df, available_bodyparts
    
    def calculate_relative_features(self, df, available_bodyparts):
        if not available_bodyparts:
            return None, 0, 0
        conf_data = []
        raw_conf_data = []
        for bp in available_bodyparts:
            conf_col = f'{bp}_confidence'
            if conf_col in df.columns:
                conf_series = df[conf_col].fillna(method='ffill').fillna(method='bfill')
                raw_conf_data.append(conf_series)
                if conf_series.max() > conf_series.min():
                    normalized = (conf_series - conf_series.min()) / (conf_series.max() - conf_series.min() + 1e-10)
                else:
                    normalized = conf_series * 0
                conf_data.append(normalized)
        if len(conf_data) < 2:
            return None, 0, 0
        conf_df = pd.concat(conf_data, axis=1)
        raw_conf_df = pd.concat(raw_conf_data, axis=1)
        avg_raw_confidence = raw_conf_df.mean().mean()
        avg_rel_confidence = conf_df.mean().mean()
        features = {}
        mean_rel_confidence = conf_df.mean(axis=1)
        features['relative_confidence'] = mean_rel_confidence
        window = int(self.fps * 2.0)
        if len(conf_df.columns) > 1:
            variance = conf_df.var(axis=1, ddof=0)
            rolling_var = variance.rolling(window=window, center=True, min_periods=1).mean()
            if rolling_var.max() > rolling_var.min():
                stability = 1 - ((rolling_var - rolling_var.min()) / (rolling_var.max() - rolling_var.min() + 1e-10))
            else:
                stability = pd.Series(0.5, index=rolling_var.index)
        else:
            rolling_std = conf_df.iloc[:, 0].rolling(window=window, center=True, min_periods=1).std()
            if rolling_std.max() > rolling_std.min():
                stability = 1 - ((rolling_std - rolling_std.min()) / (rolling_std.max() - rolling_std.min() + 1e-10))
            else:
                stability = pd.Series(0.5, index=rolling_std.index)
        features['stability'] = stability
        if len(conf_data) >= 2:
            corr_window = int(self.fps * 1.5)
            corr_scores = []
            num_pairs = min(6, len(conf_data))
            for i in range(num_pairs):
                for j in range(i+1, min(i+3, len(conf_data))):
                    corr = conf_data[i].rolling(window=corr_window, center=True, min_periods=1).corr(conf_data[j])
                    corr = corr.fillna(0).clip(0, 1)
                    corr_scores.append(corr)
            if corr_scores:
                avg_corr = pd.concat(corr_scores, axis=1).mean(axis=1)
                features['correlation'] = avg_corr.fillna(0)
        quality_score = avg_rel_confidence * 0.4 + features['stability'].mean() * 0.6
        return features, avg_raw_confidence, avg_rel_confidence, quality_score
    
    def calculate_freezing_score(self, features, sensitivity):
        if features is None:
            return pd.Series(0, index=pd.RangeIndex(1000))
        if 'stability' in features:
            freezing_score = features['stability'].copy()
        else:
            freezing_score = pd.Series(0, index=features['relative_confidence'].index)
        if 'relative_confidence' in features:
            freezing_score = freezing_score * 0.6 + features['relative_confidence'] * 0.4
        if 'correlation' in features:
            freezing_score = freezing_score * 0.7 + features['correlation'] * 0.3
        window = int(self.fps * 0.3)
        freezing_score = freezing_score.rolling(window=window, center=True, min_periods=1).mean()
        freezing_score = freezing_score.fillna(method='ffill').fillna(method='bfill').fillna(0)
        sensitivity_factor = 0.5 + (sensitivity * 0.1)
        freezing_score = freezing_score * sensitivity_factor
        freezing_score = freezing_score ** (1.0 + sensitivity * 0.1)
        return freezing_score
    
    def detect_freezing_bouts(self, freezing_score, sensitivity, avg_raw_confidence, avg_rel_confidence):
        if freezing_score.max() == 0:
            return [], {'threshold': 0.1, 'reason': 'No signal'}
        mean_score = freezing_score.mean()
        max_score = freezing_score.max()
        base_threshold = 0.15
        threshold = base_threshold * (1.0 / (1.0 + sensitivity * 0.1))
        if avg_raw_confidence < 0.05:
            threshold = threshold * 1.5
        elif avg_raw_confidence < 0.1:
            threshold = threshold * 1.2
        threshold = threshold * (1.5 - avg_rel_confidence)
        if freezing_score.max() > 0:
            if sensitivity > 3.0:
                percentile = 70
            elif sensitivity > 2.0:
                percentile = 75
            else:
                percentile = 80
            q_val = np.percentile(freezing_score[freezing_score > 0], percentile)
            threshold = max(threshold, q_val * 0.6)
            threshold = min(threshold, q_val * 0.9)
        threshold = max(threshold, mean_score * 1.2)
        threshold = max(threshold, 0.05)
        threshold = min(threshold, 0.35)
        threshold = threshold * (1.0 + (self.animals_per_video - 2) * 0.05)
        print(f"Mean score: {mean_score:.4f}, Max: {max_score:.4f}")
        print(f"Adaptive threshold: {threshold:.4f}")
        freezing_frames = freezing_score > threshold
        consistency_window = int(self.fps * 0.5)
        rolling_mean = freezing_score.rolling(window=consistency_window, center=True, min_periods=1).mean()
        consistent_frames = rolling_mean > (threshold * 0.7)
        freezing_frames = freezing_frames & consistent_frames
        changes = np.diff(np.concatenate(([0], freezing_frames.astype(int), [0])))
        bout_starts = np.where(changes == 1)[0]
        bout_ends = np.where(changes == -1)[0] - 1
        min_gap = int(self.fps * 0.8)
        merged_starts, merged_ends = [], []
        if len(bout_starts) > 0:
            current_start = bout_starts[0]
            current_end = bout_ends[0]
            for i in range(1, len(bout_starts)):
                if bout_starts[i] - current_end <= min_gap:
                    current_end = bout_ends[i]
                else:
                    merged_starts.append(current_start)
                    merged_ends.append(current_end)
                    current_start = bout_starts[i]
                    current_end = bout_ends[i]
            merged_starts.append(current_start)
            merged_ends.append(current_end)
        freezing_bouts = []
        for start, end in zip(merged_starts, merged_ends):
            duration = (end - start + 1) / self.fps
            if duration < self.min_freezing_duration or duration > self.max_freezing_duration:
                continue
            bout_score = freezing_score.iloc[start:end+1]
            mean_bout_score = bout_score.mean()
            min_bout_score = bout_score.min()
            if (mean_bout_score > threshold * 1.05 and
                min_bout_score > threshold * 0.65):
                freezing_bouts.append({
                    'start_frame': start,
                    'end_frame': end,
                    'duration_seconds': duration,
                    'mean_freezing_score': mean_bout_score,
                    'max_freezing_score': bout_score.max(),
                    'min_freezing_score': min_bout_score,
                    'start_time': start / self.fps,
                    'end_time': end / self.fps
                })
        detection_info = {
            'threshold': threshold,
            'mean_score': mean_score,
            'max_score': max_score,
            'sensitivity': sensitivity,
            'total_frames': len(freezing_score),
            'freezing_frames': freezing_frames.sum(),
            'freezing_percentage': freezing_frames.sum() / len(freezing_score) * 100
        }
        return freezing_bouts, detection_info
    
    def score_sensitivity(self, freezing_bouts, freezing_percentage, sensitivity):
        score = 0
        if freezing_percentage > 0:
            score += 10
        if 5 <= freezing_percentage <= 30:
            score += 30
        elif 1 <= freezing_percentage < 5:
            score += 15
        elif 30 < freezing_percentage <= 50:
            score += 5
        elif freezing_percentage > 50:
            score -= 15
        bout_count = len(freezing_bouts)
        if 1 <= bout_count <= 25:
            score += 15
        elif bout_count > 25:
            score -= 5
        if freezing_bouts:
            durations = [b['duration_seconds'] for b in freezing_bouts]
            avg_duration = np.mean(durations)
            if 0.8 <= avg_duration <= 10:
                score += 20
            elif avg_duration > 10:
                score += 5
        if sensitivity < 1.0 or sensitivity > 4.0:
            score -= 3
        return score
    
    def analyze_video_adaptive(self, csv_file):
        print(f"\n{'='*60}")
        print(f"ADAPTIVE ANALYSIS: {os.path.basename(csv_file)}")
        print(f"{'='*60}")
        df, available_bodyparts = self.load_data(csv_file)
        if not available_bodyparts:
            print("No body parts found!")
            return None
        if len(available_bodyparts) < 2:
            print(f"Only {len(available_bodyparts)} body parts")
        print(f"Calculating relative features...")
        features, avg_raw_confidence, avg_rel_confidence, quality_score = self.calculate_relative_features(df, available_bodyparts)
        if features is None:
            print("Could not calculate features")
            return None
        print(f"Raw confidence: {avg_raw_confidence:.6f}")
        print(f"Relative confidence: {avg_rel_confidence:.3f}")
        print(f"Quality score: {quality_score:.3f}")
        print(f"\nTesting {len(self.sensitivity_levels)} sensitivity levels:")
        sensitivity_results = []
        for sensitivity in self.sensitivity_levels:
            print(f"     Sensitivity {sensitivity:.1f}: ", end="")
            freezing_score = self.calculate_freezing_score(features, sensitivity)
            freezing_bouts, detection_info = self.detect_freezing_bouts(
                freezing_score, sensitivity, avg_raw_confidence, avg_rel_confidence
            )
            total_freezing_time = sum(b['duration_seconds'] for b in freezing_bouts)
            freezing_percentage = total_freezing_time / (len(df) / self.fps) * 100
            score = self.score_sensitivity(freezing_bouts, freezing_percentage, sensitivity)
            result = {
                'sensitivity': sensitivity,
                'freezing_percentage': freezing_percentage,
                'freezing_bouts': len(freezing_bouts),
                'total_freezing_seconds': total_freezing_time,
                'mean_freezing_score': freezing_score.mean(),
                'max_freezing_score': freezing_score.max(),
                'threshold': detection_info['threshold'],
                'score': score,
                'freezing_score': freezing_score,
                'freezing_bouts_details': freezing_bouts,
                'detection_info': detection_info
            }
            sensitivity_results.append(result)
            print(f"Freezing: {freezing_percentage:.1f}%, Score: {score:.0f}")
        best_result = max(sensitivity_results, key=lambda x: x['score'])
        best_sensitivity = best_result['sensitivity']
        print(f"\n   BEST SENSITIVITY: {best_sensitivity:.1f}")
        print(f"     Freezing: {best_result['freezing_percentage']:.1f}%")
        print(f"     Bouts: {best_result['freezing_bouts']}")
        print(f"     Score: {best_result['score']:.0f}")
        total_video_time = len(df) / self.fps
        results = {
            'video': os.path.basename(csv_file),
            'analysis_successful': True,
            'sensitivity_used': best_sensitivity,
            'video_seconds': total_video_time,
            'video_minutes': total_video_time / 60,
            'avg_raw_confidence': avg_raw_confidence,
            'avg_rel_confidence': avg_rel_confidence,
            'quality_score': quality_score,
            'freezing_bouts': best_result['freezing_bouts'],
            'total_freezing_seconds': best_result['total_freezing_seconds'],
            'freezing_percentage': best_result['freezing_percentage'],
            'mean_freezing_score': best_result['mean_freezing_score'],
            'max_freezing_score': best_result['max_freezing_score'],
            'detection_threshold': best_result['threshold'],
            'bodyparts_used': len(available_bodyparts),
            'all_sensitivities': sensitivity_results
        }
        if best_result['freezing_bouts'] > 0:
            durations = [b['duration_seconds'] for b in best_result['freezing_bouts_details']]
            results['mean_bout_duration'] = np.mean(durations)
            results['max_bout_duration'] = np.max(durations)
            results['min_bout_duration'] = np.min(durations)
            if len(best_result['freezing_bouts_details']) > 1:
                intervals = []
                for i in range(1, len(best_result['freezing_bouts_details'])):
                    interval = best_result['freezing_bouts_details'][i]['start_time'] - best_result['freezing_bouts_details'][i-1]['end_time']
                    intervals.append(interval)
                results['mean_bout_interval'] = np.mean(intervals) if intervals else 0
            else:
                results['mean_bout_interval'] = 0
        else:
            results['mean_bout_duration'] = 0
            results['max_bout_duration'] = 0
            results['min_bout_duration'] = 0
            results['mean_bout_interval'] = 0
        print(f"\nFINAL RESULTS:")
        print(f"Video: {results['video']}")
        print(f"Length: {results['video_seconds']:.0f}s")
        print(f"Animals: {self.animals_per_video}")
        print(f"Best sensitivity: {results['sensitivity_used']:.1f}")
        print(f"Freezing: {results['freezing_percentage']:.1f}% ({results['total_freezing_seconds']:.1f}s)")
        print(f"Bouts: {results['freezing_bouts']}")
        if results['freezing_bouts'] > 0:
            print(f"Avg bout: {results['mean_bout_duration']:.1f}s")
            print(f"Min bout: {results['min_bout_duration']:.1f}s")
        freezing_pct = results['freezing_percentage']
        if freezing_pct == 0:
            print(f"Interpretation: No freezing detected even with brief bout detection")
        elif freezing_pct < 5:
            print(f"Interpretation: Minimal freezing detected (brief bouts)")
        elif freezing_pct < 15:
            print(f"Interpretation: Low freezing - reasonable for restraint")
        elif freezing_pct < 30:
            print(f"Interpretation: Moderate freezing - good stress response")
        else:
            print(f"Interpretation: High freezing - strong response")
        return results
    
    def analyze_video(self, csv_file):
        return self.analyze_video_adaptive(csv_file)
    
    def run_batch_analysis(self, folder_path):
        csv_files = glob.glob(os.path.join(folder_path, "*.csv"))
        if not csv_files:
            print(f"No CSV files found in {folder_path}")
            return None
        print(f"Found {len(csv_files)} CSV files in folder")
        all_results = []
        for i, csv_file in enumerate(csv_files, 1):
            print(f"\nProcessing file {i}/{len(csv_files)}")
            try:
                results = self.analyze_video(csv_file)
                if results:
                    all_results.append(results)
                else:
                    failed_result = {
                        'video': os.path.basename(csv_file),
                        'analysis_successful': False,
                        'sensitivity_used': 0,
                        'video_seconds': 0,
                        'video_minutes': 0,
                        'avg_raw_confidence': 0,
                        'avg_rel_confidence': 0,
                        'quality_score': 0,
                        'freezing_bouts': 0,
                        'total_freezing_seconds': 0,
                        'freezing_percentage': 0,
                        'mean_freezing_score': 0,
                        'max_freezing_score': 0,
                        'detection_threshold': 0,
                        'mean_bout_duration': 0,
                        'max_bout_duration': 0,
                        'min_bout_duration': 0,
                        'bodyparts_used': 0,
                        'mean_bout_interval': 0
                    }
                    all_results.append(failed_result)
            except Exception as e:
                print(f"Error processing {csv_file}: {str(e)}")
                error_result = {
                    'video': os.path.basename(csv_file),
                    'analysis_successful': False,
                    'sensitivity_used': 0,
                    'video_seconds': 0,
                    'video_minutes': 0,
                    'avg_raw_confidence': 0,
                    'avg_rel_confidence': 0,
                    'quality_score': 0,
                    'freezing_bouts': 0,
                    'total_freezing_seconds': 0,
                    'freezing_percentage': 0,
                    'mean_freezing_score': 0,
                    'max_freezing_score': 0,
                    'detection_threshold': 0,
                    'mean_bout_duration': 0,
                    'max_bout_duration': 0,
                    'min_bout_duration': 0,
                    'bodyparts_used': 0,
                    'mean_bout_interval': 0,
                    'error': str(e)
                }
                all_results.append(error_result)
        if all_results:
            self.save_summary_to_excel(all_results, folder_path)
            return all_results
        else:
            print("No results generated")
            return None
    
    def save_summary_to_excel(self, all_results, folder_path):
        summary_data = []
        for results in all_results:
            row = {
                'Video': results['video'],
                'Analysis Successful': 'Yes' if results.get('analysis_successful') else 'No',
                'Sensitivity Used': results.get('sensitivity_used', 0),
                'Video Duration (s)': results['video_seconds'],
                'Video Duration (min)': results['video_minutes'],
                'Avg Raw Confidence': results.get('avg_raw_confidence', 0),
                'Avg Relative Confidence': results.get('avg_rel_confidence', 0),
                'Quality Score': results.get('quality_score', 0),
                'Freezing Bouts': results['freezing_bouts'],
                'Total Freezing (s)': results['total_freezing_seconds'],
                'Freezing %': results['freezing_percentage'],
                'Mean Freezing Score': results['mean_freezing_score'],
                'Max Freezing Score': results['max_freezing_score'],
                'Detection Threshold': results['detection_threshold'],
                'Mean Bout Duration (s)': results['mean_bout_duration'],
                'Max Bout Duration (s)': results['max_bout_duration'],
                'Min Bout Duration (s)': results['min_bout_duration'],
                'Body Parts Used': results['bodyparts_used'],
                'Mean Bout Interval (s)': results.get('mean_bout_interval', 0)
            }
            if 'error' in results:
                row['Error'] = results['error']
            summary_data.append(row)
        df = pd.DataFrame(summary_data)
        successful_results = [r for r in all_results if r.get('analysis_successful')]
        if successful_results:
            total_video_time = sum(r['video_seconds'] for r in successful_results)
            total_freezing_time = sum(r['total_freezing_seconds'] for r in successful_results)
            overall_freezing_pct = (total_freezing_time / total_video_time * 100) if total_video_time > 0 else 0
            videos_with_bouts = [r for r in successful_results if r['freezing_bouts'] > 0]
            stats_row = {
                'Video': 'OVERALL SUMMARY',
                'Analysis Successful': f'{len(successful_results)}/{len(all_results)}',
                'Sensitivity Used': f"Avg: {np.mean([r.get('sensitivity_used', 0) for r in successful_results]):.1f}",
                'Video Duration (s)': total_video_time,
                'Video Duration (min)': total_video_time / 60,
                'Avg Raw Confidence': np.mean([r.get('avg_raw_confidence', 0) for r in successful_results]),
                'Avg Relative Confidence': np.mean([r.get('avg_rel_confidence', 0) for r in successful_results]),
                'Quality Score': np.mean([r.get('quality_score', 0) for r in successful_results]),
                'Freezing Bouts': sum(r['freezing_bouts'] for r in successful_results),
                'Total Freezing (s)': total_freezing_time,
                'Freezing %': overall_freezing_pct,
                'Mean Freezing Score': np.mean([r['mean_freezing_score'] for r in successful_results]),
                'Max Freezing Score': np.max([r['max_freezing_score'] for r in successful_results]),
                'Detection Threshold': np.mean([r['detection_threshold'] for r in successful_results]),
                'Mean Bout Duration (s)': np.mean([r['mean_bout_duration'] for r in videos_with_bouts]) if videos_with_bouts else 0,
                'Max Bout Duration (s)': np.max([r['max_bout_duration'] for r in successful_results]),
                'Min Bout Duration (s)': np.min([r['min_bout_duration'] for r in videos_with_bouts]) if videos_with_bouts else 0,
                'Body Parts Used': np.mean([r['bodyparts_used'] for r in successful_results]),
                'Mean Bout Interval (s)': np.mean([r.get('mean_bout_interval', 0) for r in videos_with_bouts]) if videos_with_bouts else 0
            }
            stats_df = pd.DataFrame([stats_row])
            df = pd.concat([df, stats_df], ignore_index=True)
        output_file = os.path.join(folder_path, "BriefBout_Freezing_Summary.xlsx")
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Summary', index=False)
            workbook = writer.book
            worksheet = writer.sheets['Summary']
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            stats_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            if successful_results:
                last_row = len(df)
                for cell in worksheet[last_row]:
                    cell.fill = stats_fill
                    cell.font = Font(bold=True)
            for row in range(2, len(df) + 1):
                freezing_pct_cell = worksheet[f'K{row}']
                if freezing_pct_cell.value is not None and freezing_pct_cell.value != 'Freezing %':
                    try:
                        pct = float(freezing_pct_cell.value)
                        if pct > 30:
                            freezing_pct_cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                        elif pct > 15:
                            freezing_pct_cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                        elif pct >= 5:
                            freezing_pct_cell.fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")
                        elif pct > 0:
                            freezing_pct_cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    except:
                        pass
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        print(f"\n{'='*60}")
        print(f"SUMMARY SAVED TO: {output_file}")
        print(f"{'='*60}")
        if successful_results:
            print(f"\nOVERALL STATISTICS:")
            print(f"Videos analyzed: {len(successful_results)}/{len(all_results)} successful")
            print(f"Average sensitivity used: {np.mean([r.get('sensitivity_used', 0) for r in successful_results]):.1f}")
            print(f"Total video time: {total_video_time/60:.1f} minutes")
            print(f"Total freezing time: {total_freezing_time/60:.1f} minutes")
            print(f"Overall freezing: {overall_freezing_pct:.1f}%")
            print(f"Average raw confidence: {np.mean([r.get('avg_raw_confidence', 0) for r in successful_results]):.6f}")
            print(f"Average relative confidence: {np.mean([r.get('avg_rel_confidence', 0) for r in successful_results]):.3f}")
            print(f"Total freezing bouts: {sum(r['freezing_bouts'] for r in successful_results)}")
            print(f"Videos with freezing: {len(videos_with_bouts)}/{len(successful_results)}")
            if videos_with_bouts:
                print(f"Average bout duration: {np.mean([r['mean_bout_duration'] for r in videos_with_bouts]):.1f}s")
                print(f"Minimum bout duration detected: {np.min([r['min_bout_duration'] for r in videos_with_bouts if r['min_bout_duration'] > 0]):.1f}s")
        return df

if __name__ == "__main__":
    FOLDER_PATH = r"D:\Deeplabcut\RestraintStress2\HFDNaChBac\Freezing"
    ANIMALS_PER_VIDEO = 1
    print("=" * 60)
    print("ADAPTIVE FREEZING DETECTOR")
    print("=" * 60)
    print(f"KEY IMPROVEMENTS:")
    print(f"  1. MINIMUM BOUT DURATION: 0.8s ")
    print(f"  2. More adaptive detection criteria")
    print(f"  3. Shorter smoothing windows")
    print(f"Folder: {FOLDER_PATH}")
    print("=" * 60)
    detector = AdaptiveFreezingDetector(
        animals_per_video=ANIMALS_PER_VIDEO,
        fps=30
    )
    results = detector.run_batch_analysis(
        folder_path=FOLDER_PATH
    )
    print("\n" + "=" * 60)
    if results:
        successful = sum(1 for r in results if r.get('analysis_successful'))
        print(f"ANALYSIS COMPLETE!")
        print(f"Processed {len(results)} videos")
        print(f"{successful} successful analyses")
        print(f"Summary saved to: BriefBout_Freezing_Summary.xlsx")
    else:
        print("ANALYSIS FAILED")
    print("=" * 60)